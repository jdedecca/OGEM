__author__ = 'joaogorenstein'

import os, sys, csv, warnings
import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
from numpy import zeros_like, shape
import itertools as it
import openpyxl as xls
import pandas as pd
import pypsa
import igraph
import OGEM_settings
import load_flow_verification
from copy import deepcopy

def Period_Run(run_number, period, selection_data, pypsa_network, links_dataframe):
    """ Runs one expansion period. 1) Updates the network; 2) Call the expansion portfolio iteration; 3) Updates links_dataframe with selected expansion;"""

    # Creates empty selection data if first run or no candidate selected before.

    if not len(selection_data):
        selection_data = Reset_Selection_Data(pypsa_network)

    links_dataframe, pypsa_network = PyPSA_Network_Update(period, pypsa_network, links_dataframe)


    selected_links, selection_data, valid_grid, pypsa_network, links_dataframe, CBA_data, period = Expansions_Iteration(period, selection_data, links_dataframe, pypsa_network)

    Write_Output(CBA_data, period, run_number)

    # Update link status, transmission capacity, reactance and resistance.

    for link, cap in selected_links.items():

        if valid_grid.es.find(number=link)["active"] == 0:
            links_dataframe.loc[str(link), "active"] = 1
            links_dataframe.loc[str(link), "r"] = links_dataframe.loc[str(link), "r"] * links_dataframe.loc[str(link), "s_nom"] / cap
            links_dataframe.loc[str(link), "x"] = links_dataframe.loc[str(link), "x"] * links_dataframe.loc[str(link), "s_nom"] / cap
            links_dataframe.loc[str(link), "s_nom"] = cap
        else:
            links_dataframe.loc[str(link), "r"] = links_dataframe.loc[str(link), "r"] * links_dataframe.loc[str(link), "s_nom"] / (links_dataframe.loc[str(link), "s_nom"] + cap)
            links_dataframe.loc[str(link), "x"] = links_dataframe.loc[str(link), "x"] * links_dataframe.loc[str(link), "s_nom"] / (links_dataframe.loc[str(link), "s_nom"] + cap)
            links_dataframe.loc[str(link), "s_nom"] += cap

    if period == parameters["periods"] - 1:
        valid_grid = Network_Graph(period, pypsa_network, links_dataframe)
        Picture_Grid(parameters["periods"], valid_grid)

    return selected_links, selection_data, pypsa_network, links_dataframe


def Run_OPF(period, pypsa_network):
    """ Run optimal power flow for current period and all planning cases"""

    output = []

    for case in range(parameters["planning_cases"]):

        if parameters["exp_detail"] == 2:
            print("Running case %s, period %s" % (case, period))

        # Deepcopy creating temp_network, otherwise PyPSA OPF results are updated for each planning case iteration.
        temp_network = PyPSA_Time_Series(OGEM_settings.data, period, case, deepcopy(pypsa_network))

        temp_network.lines = temp_network.lines.drop(temp_network.lines.loc[temp_network.lines["s_nom"] == 0, :].index)

        temp_network.lopf(snapshots=temp_network.snapshots, verbose=False, solver_name="glpk")

        output.append(temp_network)

        if temp_network.results["Solver"][0]["Status"].key != "ok" or temp_network.results["Solver"][0]["Termination condition"].key != "optimal":
            Print_PyPSA(temp_network) # Print results for debugging.
            print("Error: Simulation did not converge or is non-optimal at period %s" % period)

    return output


def Network_Graph(period, pypsa_network, links_dataframe):
    """    # Creates a graph representing the valid options for the grid (buses and vertices) """
    # To install igraph on Windows there are two options: Search Anaconda user package databse or 2) Download wheel package from www.lfd.uci.edu/~gohlke/pythonlibs/

    valid_grid = igraph.Graph()

    valid_grid.add_vertices(parameters["buses"])

    #Create buses index that is indifferent to igraph bus order
    valid_grid.vs["pandas_index"] = pypsa_network.buses.index

    valid_grid.vs["capacity"] = 0
    valid_grid.vs["original capacity"] = 0 # Previous period capacity

    for x in valid_grid.vs:
        x["demand"] = pypsa_network.loads_t.p_set.loc[:, x["pandas_index"]].mean()
        x["type"] = pypsa_network.buses.loc[x["pandas_index"], "terminal_type"]

    for i, gen in pypsa_network.generators.iterrows():
        valid_grid.vs.find(pandas_index=gen["bus"])["capacity"] += gen["p_nom_series"][period]
        valid_grid.vs.find(pandas_index=gen["bus"])["original capacity"] += gen["p_nom_series"][max(period - 1, 0)]

    valid_grid.vs["valid"] = pypsa_network.buses["valid"]
    valid_grid.vs["xpos"] = pypsa_network.buses["x"]
    valid_grid.vs["ypos"] = pypsa_network.buses["y"]

    #Create edges composed of PyPSA lines and transport links
    edges = [[(valid_grid.vs.find(pandas_index=str(x["bus0"])).index), (valid_grid.vs.find(pandas_index=str(x["bus1"])).index)] for i, x in pypsa_network.lines.iterrows()]

    edges += [[(valid_grid.vs.find(pandas_index=str(x["bus0"])).index), (valid_grid.vs.find(pandas_index=str(x["bus1"])).index)] for i, x in pypsa_network.transport_links.iterrows()]
    valid_grid.add_edges(edges)

    valid_grid.es["number"] = [link["number"] for i, link in pypsa_network.lines.iterrows()] + [link["number"] for i, link in pypsa_network.transport_links.iterrows()]
    valid_grid.es["active"] = [link["active"] for i, link in links_dataframe.sort_values("link_type").iterrows()]
    valid_grid.es["length"] = [link["length"] for i, link in links_dataframe.sort_values("link_type").iterrows()]
    valid_grid.es["capacity"] = [link["s_nom"] for i, link in links_dataframe.sort_values("link_type").iterrows()]
    valid_grid.es["valid"] = [link["valid"] for i, link in links_dataframe.sort_values("link_type").iterrows()]
    valid_grid.es["link_type"] = [link["link_type"] for i, link in links_dataframe.sort_values("link_type").iterrows()]

    # Bus types: dummy for eventual convergence problems in pure DC networks (used for PYPOWER version, retired); owf = offshore wind farm; off = offshore hub; on, on im and on ex = average, high and low-priced buses respectively.
    # Link types: dummy connecting a dummy bus to any other bus; owf to owf connecting two owf buses; off connecting any bus to an offshore hub; radial connecting a wind farm to shore; interconnector connecting any two onshore buses;
    for x in valid_grid.es:
        types = valid_grid.vs[x.tuple]["type"]
        if "dummy" in types:
            x["type"] = "dummy"
        elif types == ["owf", "owf"]:
            x["type"] = "owf to owf"
        elif "owf" in types:
            if "off" in types:
                x["type"] = "hub"
            elif "on" in types or "on im" in types or "on ex" in types:
                x["type"] = "radial"
        elif "off" in types:
            x["type"] = "off"
        else:
            x["type"] = "interconnector"

    valid_grid.delete_edges(valid_grid.es.select(valid=0))
    valid_grid.delete_vertices(valid_grid.vs.select(valid=0))
    return valid_grid

def Picture_Grid(period, valid_grid):
    """ Create grid visualization by period in networkx, igraph not used due to compatibility issues with Windows"""

    plotperiod = deepcopy(period)
    if plotperiod == parameters["periods"]:
        period -= 1
    igraph_grid = valid_grid.subgraph(vertices=valid_grid.vs.select(type_ne="dummy").indices)
    valid_gridnx = {}
    valid_gridnx["edges"] = igraph_grid.get_edgelist()
    valid_gridnx["buses"] = igraph_grid.vcount()

    if plotperiod == parameters["periods"]:
        bus_capacity = igraph_grid.vs["capacity"]
    else:
        bus_capacity = igraph_grid.vs["original capacity"]

    valid_gridnx["labels"] = {x: 'N{0} G:{1:.0f} D:{2:.0f}\n\n'.format(igraph_grid.vs[x]["pandas_index"], bus_capacity[x], igraph_grid.vs[x]["demand"]) for x in
                              igraph_grid.vs.select(valid=1).indices}

    valid_gridnx["edge_color"] = ['0.5' for x in range(len(igraph_grid.es["valid"]))]
    valid_gridnx["line style"] = ['--' for x in range(len(igraph_grid.es["valid"]))]
    for x in igraph_grid.es.indices:
        if igraph_grid.es[x]["active"] == 1:
            if igraph_grid.es[x]["link_type"] == 0:
                valid_gridnx["edge_color"][x] = 'r'
            elif igraph_grid.es[x]["link_type"] == "line":
                valid_gridnx["edge_color"][x] = 'b'
            else:
                valid_gridnx["edge_color"][x] = 'g'
            valid_gridnx["line style"][x] = '-'

    valid_gridnx["position"] = {x: [igraph_grid.vs[x]["xpos"], igraph_grid.vs[x]["ypos"]] for x in igraph_grid.vs.select(valid=1).indices}

    valid_gridnx["bus_color"] = [None] * len(valid_gridnx["position"])
    for x in igraph_grid.vs.select(valid=1):
        if x["type"] == "owf":
            valid_gridnx["bus_color"][x.index] = 'b'
        elif x["type"] == "off":
            valid_gridnx["bus_color"][x.index] = 'g'
        elif x["type"] == "transit":
            valid_gridnx["bus_color"][x.index] = 'y'
        else:
            valid_gridnx["bus_color"][x.index] = 'r'

    if parameters["interactive"] == 0:
        plt.subplot(np.ceil(np.sqrt(parameters["periods"] + 1)), np.ceil((parameters["periods"] + 1) / np.ceil(np.sqrt(parameters["periods"] + 1))), plotperiod + 1)
    plot = plt.gca()

    plot.xaxis.set_visible(False)
    plot.yaxis.set_visible(False)

    if plotperiod == 0:
        plt.title('Initial State')
    elif plotperiod == parameters["periods"]:
        plt.title('End State - Period {:2d}'.format(period))
    else:
        plt.title('Start at Period {:2d}'.format(period))

    graph_edges = [(igraph_grid.es[x].source, igraph_grid.es[x].target,
                    {'label': "{:2d} : {:4.0f}MW".format(igraph_grid.es[x]['number'], igraph_grid.es[x]['capacity']), 'active': igraph_grid.es[x]['active']},) for x in
                   igraph_grid.es.indices]
    graph_grid = nx.empty_graph(valid_gridnx["buses"])
    graph_grid.add_edges_from(graph_edges)

    nx.draw_networkx(graph_grid, pos=valid_gridnx["position"], labels=valid_gridnx["labels"], bus_color=valid_gridnx["bus_color"], edgelist=valid_gridnx["edges"],
                     edge_color=valid_gridnx["edge_color"], style=valid_gridnx["line style"], alpha=0.5)
    nx.draw_networkx_edge_labels(graph_grid, pos=valid_gridnx["position"], edge_labels=dict([((u, v), l['label']) for u, v, l in graph_edges if l['active'] == 1]))

    plt.tight_layout()
    if parameters["interactive"] == 0:
        plt.ion()

    elif parameters["interactive"] == 1:
        plt.show()

    return


def Expansion_Portfolio(period, CBA_base, pypsa_network, valid_grid):
    """ Expansion portfolio creation following assigned rule.
    Creates portfolio of candidates and then calls function to determine the transmission capacity of each candidate.
    """

    expansions = []
    valid_expansions_dict = {}
    bus_incidence = valid_grid.get_inclist()  # Offshore buses incidence list, i.e. all the lines connected to each bus.
    active_grid = valid_grid.copy()
    active_grid.delete_edges(active_grid.es.select(active=0)) # Active grid composed only of existing lines.
    active_bus_incidence = active_grid.get_inclist()
    bus_incidence_count = np.array([x != [] for x in active_bus_incidence]) # Count of active connected lines for each bus.

    onshore_buses = pypsa_network.buses.loc[(pypsa_network.buses["terminal_type"] == "on im") | (pypsa_network.buses["terminal_type"] == "on ex") | (pypsa_network.buses["terminal_type"] == "on")]

    # Classification of high, average or low priced-buses according to average base case prices
    onshore_prices = CBA_base["bus prices"].mean(axis=0)[[int(x) for x in onshore_buses["number"]]]

    pypsa_network.buses.loc[onshore_buses.index,"terminal_type"] = "on" # Buses are presumed to be average-priced unless proven guilty.
    for i, bus in onshore_buses.iterrows():
        if np.greater(onshore_prices.max() - onshore_prices.min(), 0.1): # Price ranges must differ by more than 0.1 $/MWh to create high and low-priced buses.
            if onshore_prices[bus["number"]] == onshore_prices.max():
                pypsa_network.buses.loc[i,"terminal_type"] = "on im"
            elif onshore_prices[bus["number"]] == onshore_prices.min():
                pypsa_network.buses.loc[i,"terminal_type"] = "on ex"

    valid_grid.vs["type"] = pypsa_network.buses["terminal_type"] # Update valid grid type - pypsa_network acts as active grid.

    if parameters["expansion_mode"] == "integration_level":
        # integration_level expansion mode opposes integrated vs non-integrated typologies.

        # radial + IC typology class: offshore wind farms radial connection + all direct interconnectors expansion.
        temp_expansions = []

        for x in valid_grid.vs.select(type_in=["on im", "on ex", "on"]):  # Find shortest path among onshore buses
            for y in valid_grid.vs.select(type_in=["on im", "on ex", "on"]):
                if y != x:
                    temp_expansions += valid_grid.get_shortest_paths(x.index, to=y.index, weights="length", output="epath")

        temp_expansions = [y for x in temp_expansions for y in x]

        for x in valid_grid.vs.select(type="owf"):  # find first shortest path from owf to onshore bus
            for i, gen in pypsa_network.generators.iterrows():
                if (gen["bus"] == x["pandas_index"]) & (np.average(gen["p_nom_series"][period]) != 0):
                    min_length = min(valid_grid.es.select(bus_incidence[x.index], type="radial")["length"])
                    temp_expansions.append(valid_grid.es.select(bus_incidence[x.index], type="radial", length=min_length)[0].index)

        temp_expansions = [tuple(set(temp_expansions))]
        expansions = [[x, "radial + IC"] for x in temp_expansions] # Add candidate to portfolio.

        # radial + EXIC typology class: offshore wind farms radial connection + interconnectors between onshore buses with price differences.
        temp_expansions = []
        if valid_grid.vs.select(type="on ex").indices:
            for x in valid_grid.vs.select(type="on ex"):
                for y in valid_grid.vs.select(type_in=["on im", "on ex", "on"]):
                    if y != x:
                        temp_expansions += valid_grid.get_shortest_paths(x.index, to=y.index, weights="length", output="epath")

            temp_expansions = [y for x in temp_expansions for y in x]

            for x in valid_grid.vs.select(type="owf"):
                for i, gen in pypsa_network.generators.iterrows():
                    if (gen["bus"] == x["pandas_index"]) & (np.average(gen["p_nom_series"][period]) != 0):
                        min_length = min(valid_grid.es.select(bus_incidence[x.index], type="radial")["length"])
                        temp_expansions.append(valid_grid.es.select(bus_incidence[x.index], type="radial", length=min_length)[0].index)

            temp_expansions = [tuple(set(temp_expansions))]

            expansions += [[x, "radial + EXIC"] for x in temp_expansions] # Add candidate to portfolio.

        # owf to owf typology class: connection of two onshore buses passing through exactly two offshore wind farms.
        for z in valid_grid.vs.select(type="on ex"):
            for x in valid_grid.es.select(type="owf to owf"):

                temp_expansions = [x.index]
                on_ex_links = []
                on_ex_lengths = []
                for y in x.tuple:
                    temp_link = valid_grid.get_shortest_paths(z.index, to=y, weights="length", output="epath")[0]
                    on_ex_lengths += valid_grid.es[temp_link]["length"]
                    on_ex_links += temp_link

                for owf_index, y in enumerate(x.tuple):
                    if owf_index == np.argmin(on_ex_lengths):
                        temp_expansions.append(on_ex_links[owf_index])
                    else:
                        min_length = min(valid_grid.es.select(bus_incidence[y], type="radial")["length"])
                        temp_expansions += valid_grid.es.select(bus_incidence[y], type="radial", length=min_length).indices

            expansions += [[link, "owf to owf"] for link in tuple([temp_expansions])] # Add candidates to portfolio.

        # hub typology class: connection of all buses through one single hub bus
        for x in valid_grid.vs.select(type="off"):

            temp_on_expansions = valid_grid.es.select(_between=([x.index], valid_grid.vs.select(type_in=["on im", "on ex", "on"]).indices)).indices
            temp_owf_expansions = valid_grid.es.select(_between=([x.index], valid_grid.vs.select(type="owf").indices)).indices

            temp_on_expansions = [tuple(temp_on_expansions)]

            temp_expansions = []
            for i in range(1, len(temp_owf_expansions) + 1):
                temp_expansions += it.combinations(temp_owf_expansions, i)
            temp_owf_expansions = temp_expansions

            temp_expansions = [a + b for a in temp_owf_expansions for b in temp_on_expansions]
            expansions += [[tuple(x), "hub"] for x in temp_expansions if x != []] # Add candidates to portfolio.

        # Define total offshore power added in the period to decide on new split or split hybrid expansion candidates.
        total_offshore_p_nom = pypsa_network.buses[pypsa_network.buses["terminal_type"] == "owf"]["p_nom"] - \
                                   (pypsa_network.buses["previous_p_nom"] * bus_incidence_count)[
                                       pypsa_network.buses["terminal_type"] == "owf"]

        # Split and hybrid split typology classes: connection of onshore buses through single wind farms or hybridization with direct interconnection
        # Create set of onshore buses to connect through split typologies
        split_buses = set()
        for y in valid_grid.vs.select(type="on ex"):
            for z in valid_grid.vs.select(type_in=["on", "on im"]):
                split_buses = split_buses | {(y.index, z.index)}
        for y in valid_grid.vs.select(type="on"):
            for z in valid_grid.vs.select(type="on im"):
                split_buses = split_buses | {(y.index, z.index)}

        # Combine existing onshore bus pairs for the number of existing wind farms
        split_plans = [x for x in it.combinations(split_buses, len(valid_grid.vs.select(type="owf")))]
        split_farms = valid_grid.vs.select(type="owf").indices
        reversed_split_farms = list(reversed(split_farms))

        # Create the hybrid expansion candidates using the onshore bus pairs starting with reverse pair order
        for split_connection in split_plans:
            split_expansion = []

            # If the new capacity of the wind farm > 0 connect the onshore buses through the wind farm
            for i, x in enumerate(split_connection[:-1]): #Leave one onshore pair for direct interconnection.
                if total_offshore_p_nom[str(split_farms[i])] != 0:
                    split_expansion += [valid_grid.es.find(_between=([split_farms[i]], [x[0]])).index, valid_grid.es.find(_between=([split_farms[i]], [x[1]])).index]

            # If the new capacity of the wind farm == 0 connect the onshore buses directly
            for i, x in enumerate(split_connection[:-1]): #Leave one onshore pair for direct interconnection.
                if total_offshore_p_nom[str(split_farms[i])] == 0:
                    split_expansion += [valid_grid.es.find(_between=([x[0]], [x[1]])).index]

            # Directly interconnect the lef-out pair.
            split_expansion += [valid_grid.es.find(_between=([split_connection[-1][0]], [split_connection[-1][1]])).index]

            # Radially connect the left-out wind farm if new capacity > 0
            if total_offshore_p_nom[str(split_farms[-1])] != 0:
                min_length = min(valid_grid.es.select(bus_incidence[split_farms[-1]], type="radial")["length"])
                split_expansion += [valid_grid.es.select(bus_incidence[split_farms[-1]], type="radial", length=min_length)[0].index]

            expansions += [[tuple(split_expansion), "hybrid split"]] # Add candidate to portfolio.

            # Hybrid splits with reversed wind farm order.
            split_expansion = []
            for i, x in enumerate(split_connection[:-1]):
                if total_offshore_p_nom[str(split_farms[-i - 1])] != 0:
                    split_expansion += [valid_grid.es.find(_between=([split_farms[-i - 1]], [x[0]])).index,
                                        valid_grid.es.find(_between=([split_farms[-i - 1]], [x[1]])).index]
            for i, x in enumerate(split_connection[:-1]):
                if total_offshore_p_nom[str(split_farms[-i - 1])] == 0:
                    split_expansion += [valid_grid.es.find(_between=([x[0]], [x[1]])).index]

            split_expansion += [valid_grid.es.find(_between=([split_connection[-1][0]], [split_connection[-1][1]])).index]
            if total_offshore_p_nom[str(split_farms[0])] != 0:
                min_length = min(valid_grid.es.select(bus_incidence[split_farms[0]], type="radial")["length"])
                split_expansion += [
                    valid_grid.es.select(bus_incidence[split_farms[0]], type="radial", length=min_length)[0].index]

            expansions += [[tuple(split_expansion), "hybrid split"]] # Add candidate to portfolio.

        split_plans += [(x, x) for x in split_buses]


        # Create the pure split expansion candidates.
        for split_connection in split_plans:

            # Direct order split candidate creation.
            split_expansion = []
            for i, x in enumerate(split_connection):
                if total_offshore_p_nom[str(split_farms[i])] != 0:
                    split_expansion += [valid_grid.es.find(_between=([split_farms[i]], [x[0]])).index, valid_grid.es.find(_between=([split_farms[i]], [x[1]])).index]
            for i, x in enumerate(split_connection):
                if total_offshore_p_nom[str(split_farms[i])] == 0:
                    split_expansion += [valid_grid.es.find(_between=([x[0]], [x[1]])).index]
            expansions += [[tuple(split_expansion),"split plan"]]

            # Reversed order split candidate creation.
            split_expansion = []
            for i, x in enumerate(split_connection):
                if total_offshore_p_nom[str(reversed_split_farms[i])] != 0:
                    split_expansion += [valid_grid.es.find(_between=([reversed_split_farms[i]], [x[0]])).index,
                                        valid_grid.es.find(_between=([reversed_split_farms[i]], [x[1]])).index]
            for i, x in enumerate(split_connection):
                if total_offshore_p_nom[str(reversed_split_farms[i])] == 0:
                    split_expansion += [valid_grid.es.find(_between=([x[0]], [x[1]])).index]
            expansions += [[tuple(split_expansion), "split plan"]]


    else:
        expansions = [(x,) for x in valid_grid.es.indices]

    valid_expansions = []

    # Determine capacity for each candidate, varying according to the wind farm and interconnector capacity parameters.
    for x in expansions:

        for wind_sens in list(map(float, str(parameters["wind_cap_sens"]).split())):
            for IC_cap in list(map(float, str(parameters["IC_cap"]).split())):

                temp_expansion = Candidate_Capacity(x[0], x[1], wind_sens, pypsa_network, valid_grid, bus_incidence_count, IC_cap) + [wind_sens]
                if sum(temp_expansion[1]) > 0:
                    valid_expansions.append(temp_expansion + [wind_sens])

    valid_expansions_temp = []

    # Variate candidate portfolio according to the overall capacity sensitivity parameter.
    for x in list(map(float, str(parameters["capacity_sens"]).split())):
        for y in valid_expansions:

            links, igraph_links, cap = zip(*sorted(zip(valid_grid.es[y[0]]["number"], y[0], [z * x for z in y[1]])))
            valid_expansions_temp.append([links, igraph_links, cap, y[2], y[3], y[4], y[5], x])

    valid_expansions = deepcopy(valid_expansions_temp)
    return valid_expansions

def Candidate_Capacity(exp, type, wind_sens, pypsa_network, valid_grid, bus_incidence_count, IC_cap):
    """ Defines the candidate capacity by typology """

    #  Retrieval of availability factors.
    columns = OGEM_settings.data["RES_availability"][0][1:] # The columns header contain the names of the technologies.
    RES_availability = np.array(OGEM_settings.data["RES_availability"][1:], dtype=float)
    RES_availability = pd.DataFrame(RES_availability[:, 1:], index=RES_availability[:, 0], columns=columns)

    # The average offshore wind availability is used in the definition of capacities of offshore wind lines.
    wind_avg = RES_availability.loc[:, "off_wind"].mean()

    capacity = [IC_cap] * len(exp)
    buses = list(set(it.chain(*[valid_grid.vs[edge.tuple].indices for edge in valid_grid.es[exp]])))
    buses = valid_grid.vs[buses]["pandas_index"] # Buses belonging to this candidate.
    im_count = pypsa_network.buses.loc[buses, "terminal_type"].value_counts()[["on im", "on"]].sum() # Number of non-low priced buses for dividing the hub exports.
    zones = set(list([z for y in buses for z in pypsa_network.buses.loc[y, "tso"]])) # TSO zones belonging to this candidate.

    z = pypsa_network.buses.loc[buses, :] # Pandas buses for this candidate.

    # Hub capacity is the new or non-connected capacity of adjusted-offshore wind and low-priced onshore buses
    hub_cap = (z[z["terminal_type"] == "owf"]["p_nom"].sum() - (z["previous_p_nom"] * bus_incidence_count[[int(x) for x in buses]])[
        z["terminal_type"] == "owf"].sum()) * wind_avg * wind_sens + z[z["terminal_type"] == "on ex"]["p_nom"].count() * IC_cap
    total_offshore_p_nom = z[z["terminal_type"] == "owf"]["p_nom"] - (z["previous_p_nom"] * bus_incidence_count[[int(x) for x in buses]])[z["terminal_type"] == "owf"]

    #
    for j, y in enumerate(exp):
        link_buses_index = [str(i) for i in valid_grid.es[y].tuple]
        link_buses = pypsa_network.buses.loc[link_buses_index, :] # Bus pair to line in question.

        # owf to owf capacity is cumulative from low to high-priced nodes, summing the adjusted new wind farm capacities.
        if type == "owf to owf":
            if list(link_buses["terminal_type"]) == ["owf", "owf"]:
                capacity[j] = IC_cap + total_offshore_p_nom.min() * wind_avg * wind_sens
            elif "on ex" not in list(link_buses["terminal_type"]):
                capacity[j] = IC_cap + total_offshore_p_nom.sum() * wind_avg * wind_sens

        # radial capacities are standard for interconnectors and adjusted new capacity for wind farms
        elif type in ["radial + IC", "radial + EXIC"]:
            if "owf" in list(link_buses["terminal_type"]):
                capacity[j] = total_offshore_p_nom[link_buses_index].sum() * wind_sens

        # Split capacities are cumulative from low to high priced nodes, summing the adjusted new single wind farm capacity.
        elif type == "split plan":
            if "owf" not in list(link_buses["terminal_type"]):
                capacity[j] = IC_cap
            elif not j%2:
                capacity[j] = IC_cap
            else:
                capacity[j] = IC_cap + total_offshore_p_nom[link_buses_index].sum() * wind_avg * wind_sens

        # Hybrid split capacities are standard for direct interconnection or cumulative from low to high priced nodes, summing the adjusted new single wind farm capacity for the split part of the candidate.
        elif type == "hybrid split":
            if "owf" not in list(link_buses["terminal_type"]):
                capacity[j] = IC_cap
            elif ("owf" in list(link_buses["terminal_type"])) & (j == (len(exp) - 1)):
                capacity[j] = total_offshore_p_nom[link_buses_index].sum() * wind_avg * wind_sens
            elif not j % 2:
                capacity[j] = IC_cap
            else:
                capacity[j] = IC_cap + total_offshore_p_nom[link_buses_index].sum() * wind_avg * wind_sens
        elif type in ["hub"]:
            if "on" in list(link_buses["terminal_type"]) or "on im" in list(link_buses["terminal_type"]):
                capacity[j] = hub_cap / im_count
            elif "owf" in list(link_buses["terminal_type"]):
                capacity[j] = total_offshore_p_nom[link_buses_index].sum() * wind_avg * wind_sens
        else:
            print("Error: Expansion candidate type not recognized for capacity determination")
            sys.exit()

    # Eliminate lines with capacity == 0.
    exp, capacity = [[exp[i] for i, cap in enumerate(capacity) if cap > 0], [capacity[i] for i, cap in enumerate(capacity) if cap > 0]]
    buses = list(set(it.chain(*[valid_grid.vs[edge.tuple].indices for edge in valid_grid.es[exp]])))
    buses = valid_grid.vs[buses]["pandas_index"]
    return [exp, capacity, buses, zones, type]


def Output_Recovery(CBA_data, exp_ind, output, links_dataframe):
    """ Recovers OPF data:
        Total operational cost
        Nodal prices and demand
        Generator dispatch and cost
        Lines flow and price difference
    """

    bus_shadow = np.zeros((parameters["planning_cases"], parameters["case_segments"], parameters["buses"]))
    bus_demand = bus_shadow.copy()
    branches_flow = np.zeros((parameters["planning_cases"], parameters["case_segments"], parameters["number of links"]))
    branches_prices_delta = branches_flow.copy()

    ofv = [segment.objective for segment in output]
    generator_gen = np.array([np.sum(pd.concat([segment.generators_t.p, segment.storage_units_t.p], axis=1), axis=0) for segment in output])
    generator_cost = [list(
        pd.concat([segment.generators_t.p, segment.storage_units_t.p], axis=1).sum(axis=0) * pd.concat([segment.generators.marginal_cost, segment.storage_units.marginal_cost],
                                                                                                       axis=0)) for segment in output]
    bus_shadow = np.array([np.divide(np.array(segment.buses_t.marginal_price), np.array(segment.snapshot_weightings)[:, None]) for segment in output])
    bus_demand = [list(np.sum(segment.loads_t.p.reindex(columns=segment.buses.index, fill_value=0), axis=0)) for segment in output]

    for case, case_network in enumerate(output):
        branches_flow[case, :, case_network.lines.index.astype(int)] = np.abs(case_network.lines_t.p0[case_network.lines.index]).transpose()
        branches_flow[case, :, case_network.transport_links.index.astype(int)] += np.abs(case_network.transport_links_t.p0[case_network.transport_links.index]).transpose()


    branches_flow = np.average(branches_flow, axis=1, weights=parameters["segments_probability"])
    bus_shadow = np.average(bus_shadow, axis=1, weights=parameters["segments_probability"])

    branches_prices_delta = np.zeros((parameters["planning_cases"], parameters["number of links"]))
    branches_prices_delta[:, links_dataframe["number"]] = bus_shadow[:, links_dataframe["bus0"]] - bus_shadow[:, links_dataframe["bus1"]]
    branches_prices_delta = np.abs(branches_prices_delta)

    index = [(exp_ind, case) for case in range(parameters["planning_cases"])]
    CBA_data["generator gen"] = CBA_data["generator gen"].append(pd.DataFrame(generator_gen, index=index, columns=gen_sto_columns))
    CBA_data["generator cost"] = CBA_data["generator cost"].append(pd.DataFrame(generator_cost, index=index, columns=gen_sto_columns))
    CBA_data["bus prices"] = CBA_data["bus prices"].append(pd.DataFrame(bus_shadow, index=index, columns=bus_columns))
    CBA_data["bus demand"] = CBA_data["bus demand"].append(pd.DataFrame(bus_demand, index=index, columns=bus_columns))
    CBA_data["link flow"] = CBA_data["link flow"].append(pd.DataFrame(branches_flow, index=index, columns=link_columns))
    CBA_data["link price delta"] = CBA_data["link price delta"].append(pd.DataFrame(branches_prices_delta, index=index, columns=link_columns))
    CBA_data["op cost"] = CBA_data["op cost"].append(pd.DataFrame(ofv, index=index, columns=["op cost"]))

    return CBA_data


def Onshore_Distribution(CBA_data, valid_grid, pypsa_network, base_flag=0):
    """ Redistribute benefits and costs according to selected algorithm, reallocating any offshore bus welfare to onshore """

    CBA_data["ons nodal producer surplus"] = zeros_like(CBA_data["nodal producer surplus"])
    CBA_data["ons nodal benefit"] = CBA_data["nodal benefit"].copy()
    CBA_data["ons nodal congestion rent"] = CBA_data["nodal benefit"].copy()


    bus_incidence = valid_grid.get_inclist()  # Offshore buses incidence list, i.e. all the lines connected to each bus.
    CBA_data["onshore distribution"] = np.zeros(
        (len(CBA_data["clusters"]), parameters["buses"], parameters["buses"]))

    # Distribute benefits and costs inside cluster only.
    for i, x in enumerate(CBA_data["clusters"]):
        onshore_buses = None
        for y in x:
            for z in y:
                if CBA_data["valid expansions"].ix[i, "Type"] in ["hub", "on ex split", "on split", "split plan", "hybrid split", "owf to owf"]:  # hub and owf costs in integrated typologies are shared among onshore buses
                    tso = valid_grid.vs.select(type_in=["on", "on ex", "on im"])["pandas_index"]
                elif CBA_data["valid expansions"].ix[i, "Type"] in ["radial", "radial + IC", "radial + EXIC", "None"]:  # wind farm costs in radial typologies go to their tso bus
                    tso = pypsa_network.buses.loc[z, "tso"]
                else:
                    print("Error: Expansion candidate type not recognized for onshore distribution", CBA_data["valid expansions"].ix[i, "Type"])
                    sys.exit()
                onshore_buses = [int(bus) for bus in set(y).intersection(tso, CBA_data["valid expansions"].ix[i, "Buses"])] # Onshore buses sharing in costs and benefits must pertain to current candidate.

                if pypsa_network.buses.loc[z, "terminal_type"] in ["owf", "off"]:
                    if onshore_buses:
                        CBA_data["onshore distribution"][i, int(z), onshore_buses] = 1 / len(onshore_buses) # Wind farm costs and benefits sharing.

        all_onshore = pypsa_network.buses.loc[
            (pypsa_network.buses["terminal_type"] == "on im") | (pypsa_network.buses["terminal_type"] == "on ex") | (
                pypsa_network.buses["terminal_type"] == "on")]["number"]

        for y in all_onshore:
            if CBA_data["onshore distribution"][i, y, y] == 0:
                CBA_data["onshore distribution"][i, y, y] = 1 # An onshore bus must receive its own cost and benefits.

        # Share producer surplus, total nodal benefit and congestion rent differences.
        CBA_data["ons nodal producer surplus"][i] = np.dot(CBA_data["nodal producer surplus"].loc[i], CBA_data["onshore distribution"][i])
        CBA_data["ons nodal benefit"].loc[i, :] = np.dot(CBA_data["nodal benefit"].loc[i], CBA_data["onshore distribution"][i])
        CBA_data["ons nodal congestion rent"].loc[i, :] = np.dot(CBA_data["nodal congestion rent"][i], CBA_data["onshore distribution"][i])

    CBA_data["producer surplus"] = CBA_data["consumer payments"] - np.sum(CBA_data["ons nodal congestion rent"], axis=1) - CBA_data["total op cost"]

    CBA_data["benefits"] = CBA_data["producer surplus"] + np.sum(CBA_data["ons nodal congestion rent"], axis=1) - CBA_data["consumer payments"]

    return CBA_data

def Welfare_Calculation(CBA_data, pypsa_network, base_flag=0):
    """ Calculates nodal welfare components """

    # Define weighing of cases equally.
    cases_probability = np.repeat(parameters["cases_probability"] / np.sum(parameters["cases_probability"]), len(CBA_data["op cost"].index.levels[0]))

    # Calculate non-aggregate welfare components, where offshore buses have welfare != 0.
    CBA_data["total op cost"] = CBA_data["op cost"].loc[:, "op cost"].mul(cases_probability, axis=0).sum(level="expansion")
    CBA_data["consumer payments"] = (CBA_data["bus demand"] * CBA_data["bus prices"]).mul(cases_probability, axis=0).sum(level="expansion").sum(axis=1)
    CBA_data["congestion rent"] = (CBA_data["link flow"] * CBA_data["link price delta"]).mul(cases_probability, axis=0).sum(level="expansion").sum(axis=1)
    CBA_data["nodal payments"] = (CBA_data["bus demand"] * CBA_data["bus prices"]).mul(cases_probability, axis=0).sum(level="expansion")
    CBA_data["branch congestion rent"] = (CBA_data["link flow"] * CBA_data["link price delta"]).mul(cases_probability, axis=0).sum(level="expansion")
    CBA_data["nodal prices"] = CBA_data["bus prices"].mul(cases_probability, axis=0).sum(level="expansion")


    # Build congestion distribution matrix dividing congestion rents equally among the two line buses.
    CBA_data["congestion distribution"] = np.zeros((parameters["number of links"], parameters["buses"]))
    for i, link in pypsa_network.lines.iterrows():
        CBA_data["congestion distribution"][link["number"], int(link["bus0"])] = 0.5
        CBA_data["congestion distribution"][link["number"], int(link["bus1"])] = 0.5

    for i, link in pypsa_network.transport_links.iterrows():
        CBA_data["congestion distribution"][link["number"], int(link["bus0"])] = 0.5
        CBA_data["congestion distribution"][link["number"], int(link["bus1"])] = 0.5

    # Aggregate congestion rent per bus.
    CBA_data["nodal congestion rent"] = np.dot(CBA_data["branch congestion rent"], CBA_data["congestion distribution"])

    # Variable declations.
    CBA_data["nodal gen cost"] = CBA_data["bus demand"].copy()
    CBA_data["nodal generation"] = CBA_data["nodal gen cost"].copy()
    CBA_data["nodal gen cost"].loc[:, :] = 0
    CBA_data["nodal generation"].loc[:, :] = 0

    # Aggregate generation cost and dispatch per bus.

    for i, gen in pypsa_network.generators.iterrows():
        CBA_data["nodal gen cost"].loc[:, int(gen["bus"])] += CBA_data["generator cost"].loc[:, gen["name"]]
        CBA_data["nodal generation"].loc[:, int(gen["bus"])] += CBA_data["generator gen"].loc[:, gen["name"]]

    for i, sto in pypsa_network.storage_units.iterrows():
        CBA_data["nodal gen cost"].loc[:, int(sto["bus"])] += CBA_data["generator cost"].loc[:, sto["name"]]
        CBA_data["nodal generation"].loc[:, int(sto["bus"])] += CBA_data["generator gen"].loc[:, sto["name"]]

    CBA_data["nodal producer surplus"] = (CBA_data["nodal generation"] * CBA_data["bus prices"]).mul(cases_probability, axis=0).sum(level="expansion") - CBA_data[
        "nodal gen cost"].mul(cases_probability, axis=0).sum(level="expansion")
    CBA_data["nodal benefit"] = CBA_data["nodal producer surplus"] + CBA_data["nodal congestion rent"] - CBA_data["nodal payments"]

    return CBA_data


def Cost_Distribution(CBA_data, CBA_base):
    """ Future implementation of alternative investment costs redistribution mechanisms.
        Other mechanisms than onshore redistribution
    """
    np.seterr(divide='ignore', invalid='ignore')  # Overrides numpy error warning for division for buses with cost = 0

    if parameters["redistribution_mechanism"] == "onshore redistribution":
        # Defines nodal benefit to cost ratio with only a redistribution to onshore buses
        for i, x in CBA_data["nodal inv cost"].iterrows():
            CBA_data["dist nodal inv cost"].loc[i, :] = np.dot(x, CBA_data["onshore distribution"][i])

    elif parameters["redistribution_mechanism"] == "cost socialization":

        CBA_data["dist nodal inv cost"] = np.divide(CBA_data["cost"], CBA_data["on_bus_count"])[:, None] * CBA_data["on_bus_matrix"]

    elif parameters["redistribution_mechanism"] == "egalitarian non-separable cost":
        # For expansion cases, calculate non-separable cost and allocate equally among onshore buses participating in the specific expansion

        CBA_data["non-separable cost"] = CBA_data["cost"] - np.sum(CBA_data["nodal sep inv cost"], axis=1)
        CBA_data["SC nodal inv cost"] = CBA_data["nodal sep inv cost"] + np.divide(CBA_data["non-separable cost"], CBA_data["on_bus_count"])[:, None] * CBA_data[
            "on_bus_matrix"]  # Nodal costs = separable costs + non-separable costs divided among all expansion onshore buses

        # Calculate radial cost for buses for radial expansion
        for i, x in CBA_data["valid expansions"].iterrrows():
            if x["Type"] == "radial + IC":
                radial_index = i

        CBA_data["nodal radial cost"] = CBA_data["nodal inv cost"][radial_index]

        for i, x in CBA_data["SC nodal inv cost"].iterrows():
            CBA_data["dist nodal inv cost"].loc[i, :] = np.dot(CBA_data["SC nodal inv cost"].loc[i, :], CBA_data["onshore distribution"][i])

    elif parameters["redistribution_mechanism"] == "host compensation by positive net benefit":

        CBA_data["diff net nodal benefit"] = (CBA_data["nodal benefit"] - CBA_base["nodal benefit"][0]) * parameters["ben_multiplier"] - CBA_data["nodal inv cost"] * parameters[
            "cost_multiplier"]
        total_compensation = CBA_data["cost"].copy()

        bearers_beneficiaries = zeros_like(CBA_data["diff net nodal benefit"])
        for i, x in CBA_data["valid expansions"].iterrows():
            bearers_beneficiaries[i, x["buses"]] = (CBA_data["diff net nodal benefit"][i, x["buses"]] < 0) * -1
            bearers_beneficiaries[i, CBA_data["diff net nodal benefit"][i] > 0] = 1

        total_compensation = np.sum(CBA_data["diff net nodal benefit"] * (bearers_beneficiaries == -1), axis=1)
        total_net_benefits = np.sum(CBA_data["diff net nodal benefit"] * (bearers_beneficiaries == 1), axis=1)
        compensation_adjustment = np.minimum(-total_compensation, total_net_benefits) / total_compensation * -1

        CBA_data["dist nodal inv cost"] = CBA_data["nodal inv cost"] + (CBA_data["diff net nodal benefit"] * (bearers_beneficiaries == -1) * compensation_adjustment[:, None]) / \
                                                                       parameters["cost_multiplier"]

        CBA_data["dist nodal inv cost"] = CBA_data["dist nodal inv cost"] + np.nan_to_num(
            (CBA_data["diff net nodal benefit"] * (bearers_beneficiaries == 1)) / total_net_benefits[:, None] * (total_compensation * compensation_adjustment * -1)[:, None]) / \
                                                                            parameters["cost_multiplier"]

    else:
        print("Redistribution mechanism not recognized")
        sys.exit()

    # Only onshore nodes must have final welfare components.
    for i, x in CBA_data["dist nodal inv cost"].iterrows():
        x = np.dot(x, CBA_data["onshore distribution"][i])

    return CBA_data


def Cost_Calculation(CBA_data, valid_grid, pypsa_network):
    """ Calculates total and separable costs used in the welfare analysis """

    onshore_buses = pypsa_network.buses["terminal_type"].isin(["on ex", "on", "on im"]).tolist() # All system onshore buses.

    for index, expansion in CBA_data["valid expansions"].iterrows():
        for bus in expansion["Buses"]:
            CBA_data["on_bus_matrix"].loc[index, int(bus)] = True
    CBA_data["on_bus_matrix"] = CBA_data["on_bus_matrix"] & onshore_buses # boolean matrix with onshore buses pertaining to candidate.

    CBA_data["on_bus_count"] = CBA_data["on_bus_matrix"].sum(axis = 1)

    #
    for exp_ind, [pandas_number, expansion, cap, buses, zones, type, wind_sens, capacity_sens, links_string, capacity_string] in CBA_data["valid expansions"].iterrows():
        expansion_buses = dict()

        # Creation of dictionary with the expansion buses
        for x in valid_grid.es[expansion]:
            expansion_buses.update({x.source: [0, 0, 0, 0, 0, []]})  # Format is bus :[point-to-point terminals, PtP capacity, multiterminals, MT capacity, cost, connected lines]
            expansion_buses.update({x.target: [0, 0, 0, 0, 0, []]})

        # Updates all buses of the dictionary with their connected links.
        for x in valid_grid.es[expansion]:
            expansion_buses[x.source][5].append(x.index)
            expansion_buses[x.target][5].append(x.index)

        # Calculate candidate expansion cost, nodal cost and length.
        CBA_data["cost"].loc[exp_ind, "cost"], CBA_data["length"].loc[exp_ind, "length"], CBA_data["nodal inv cost"].loc[exp_ind, :] = Candidate_Cost_Calculation(exp_ind,
                                                                                                                                                                  expansion, cap,
                                                                                                                                                                  buses, zones,
                                                                                                                                                                  expansion_buses,
                                                                                                                                                                  type, valid_grid,
                                                                                                                                                                  pypsa_network)

        # Calculate separable cost for each bus.
        for x, y in expansion_buses.items():

            removal_vector = np.array([False * parameters["number of expansions"]])
            # Vector with buses which should be removed to calculate separable cost.
            for i in y[5]:
                removal_vector = removal_vector | [j == i for j in expansion]

            # Remaining cost calculation after the bus is removed.
            CBA_data["marginal nodal cost"].loc[exp_ind, x] = Candidate_Cost_Calculation(exp_ind, [expansion[i] for i in range(0, len(expansion)) if removal_vector[i] == 0],
                                                                                         [cap[i] for i in range(0, len(cap)) if removal_vector[i] == 0], buses, zones,
                                                                                         expansion_buses, type, valid_grid, pypsa_network)[
                0]
            # Separable cost = total cost - (remaining cost without bus / number of candidate onshore buses)
            CBA_data["nodal sep inv cost"].loc[exp_ind, x] = (CBA_data["cost"].loc[exp_ind, "cost"] - CBA_data["marginal nodal cost"].loc[exp_ind, x]) / CBA_data["on_bus_count"][
                exp_ind]
    return CBA_data

def Candidate_Cost_Calculation(exp_ind, expansion, cap, buses, zones, expansion_buses, type, valid_grid, pypsa_network):
    """ Calculates the investment cost for any given candidate"""

    cost = 0
    length = 0
    cost_vector = np.zeros(parameters["buses"])
    # Resets for each bus the vector [point-to-point terminals, PtP capacity, multiterminals, MT capacity]
    for x, y in expansion_buses.items():
        y[0:4] = [0, 0, 0, 0]

    for lin_exp, lin_cap in zip(expansion, cap):

        # Adds candidate line cost proportional to length and capacity increase
        lin_length = pypsa_network.lines.loc[pypsa_network.lines["number"] == valid_grid.es[lin_exp]['number'], "length"].sum() + pypsa_network.transport_links.loc[
            pypsa_network.transport_links["number"] == valid_grid.es[lin_exp]['number'], "length"].sum()
        length += lin_length
        lin_cost = lin_length * lin_cap * parameters["cap_length"] + lin_length * parameters["length"] + lin_cap * parameters["cap"]
        cost += lin_cost
        cost_vector[list(valid_grid.es[lin_exp].tuple)] += lin_cost / 2 # Allocates half of each line cost to each bus

        # Calculates expansion_buses = [point-to-point terminals, PtP capacity, multiterminals, MT capacity]
        for y in valid_grid.es[lin_exp].tuple:
            # point-to-point lines always pay their full capacity
            if valid_grid.es[lin_exp]["link_type"] == "link":
                expansion_buses[y][1] += lin_cap
                expansion_buses[y][0] += 1
            # multiterminal lines cost follow specific rules per terminal type (null, max or sum)
            elif valid_grid.es[lin_exp]["link_type"] == "line":
                if valid_grid.vs[y]["type"] in ["on", "on ex", "on im"]:
                    expansion_buses[y][3] += lin_cap
                    expansion_buses[y][2] += 1
                elif valid_grid.vs[y]["type"] == "off":
                    expansion_buses[y][3] += 0
                    expansion_buses[y][2] += 1
                elif valid_grid.vs[y]["type"] == "owf":
                    if type in ["owf to owf", "on ex split", "on split", "split plan", "hybrid split"]:
                        expansion_buses[y][3] = max(expansion_buses[y][3], lin_cap)
                        expansion_buses[y][2] += 1
                    elif type in ["radial + IC", "radial + EXIC", "hub"]:
                        #print("Reconsider")
                        expansion_buses[y][3] += lin_cap
                        expansion_buses[y][2] += 1
                    else:
                        print("Error: Expansion type not found")
                        sys.exit()
            else:
                print("Error: Link type not found")
                sys.exit()


    # Calculates nodal costs proportional to number of lines and capacity, and terminal unit costs
    for x, y in expansion_buses.items():
        if valid_grid.vs[x]["type"] == "off":
            bus_cost = parameters["off_tc"] * (y[1] + y[3])
        elif valid_grid.vs[x]["type"] == "owf":
            bus_cost = parameters["owf_tc"] * (y[1] + y[3])
        elif valid_grid.vs[x]["type"] in ["on", "on im", "on ex"]:
            bus_cost = parameters["on_tc"] * (y[1] + y[3])
        else:
            print("Error: Bus terminal cost not found")
            sys.exit()

        cost += bus_cost
        cost_vector[x] += bus_cost

    return cost, length, cost_vector


def Expansion_Selection(CBA_base, CBA_data):
    """ Selects candidate according to ratio or absoluted net welfare change and to social/regional/regional pareto scope """

    selection_vector = np.zeros(shape=shape(CBA_data["cost"])) # Boolean vector with selected candidate == TRUE
    selected_links = {}

    # Benefit types calculation.
    CBA_data["ons relative nodal benefit"] = CBA_data["ons nodal benefit"].sub(CBA_base["ons nodal benefit"].loc[0, :])
    CBA_data["net social cost benefit"] = - (CBA_data["total op cost"].sub(CBA_base["total op cost"].loc[0])) * parameters["ben_multiplier"] - CBA_data["cost"].loc[:, "cost"] * \
                                                                                                                                               parameters["cost_multiplier"] # Social net welfare change.

    CBA_data["dist net nodal benefit"] = CBA_data["ons relative nodal benefit"] * parameters["ben_multiplier"] - CBA_data["dist nodal inv cost"] * parameters["cost_multiplier"]

    for i, x in CBA_data["valid expansions"].iterrows():
        buses = list(map(float, x["Buses"]))
        CBA_data["bilateral benefit"].loc[i] = CBA_data["dist net nodal benefit"].loc[i, buses].sum() # Regional net welfare change (only for candidate buses.
        if all(CBA_data["dist net nodal benefit"].loc[i, buses] >= -0.01):
            CBA_data["unilateral benefit"].loc[i] = CBA_data["dist net nodal benefit"].loc[i, buses].sum() # Regional pareto net welfare change (only for candidate buses)

    # Benefit mode:0 = social cost; 1 = social welfare; 2 = regional welfare; 3 = regional pareto welfare;
    if parameters["benefit_type"] == 0:
        CBA_data["bcd"] = CBA_data["net social cost benefit"]
    elif parameters["benefit_type"] == 1:
        CBA_data["bcd"] = np.sum(CBA_data["dist net nodal benefit"], axis=1)
    elif parameters["benefit_type"] == 2:
        CBA_data["bcd"] = CBA_data["bilateral benefit"]
    elif parameters["benefit_type"] == 3:
        CBA_data["bcd"] = CBA_data["unilateral benefit"]
    else:
        print("Error: benefit type does not exist")
        sys.exit()

    # Net welfare ratio calculation.
    CBA_data["bcr"] = CBA_data["bcd"] / CBA_data["cost"].loc[:, "cost"] / parameters["cost_multiplier"]

    stop_criterion = parameters["stop_threshold"]  # Stop criterion determines lower acceptable threshold for net welfare ratio, otherwise no candidate is selected.

    if parameters["selection_criterion"] == 0:  # Net welfare ratio used for selection.
        max_indicator = max(CBA_data["bcr"])
        selection_vector = np.logical_and(np.greater_equal(CBA_data["bcr"], parameters["selection_threshold"] * max_indicator), np.greater(CBA_data["bcr"], stop_criterion))
    elif parameters["selection_criterion"] == 1:  # Absolute net welfare used for selection.
        max_indicator = max(CBA_data["bcd"])
        selection_vector = np.logical_and(np.greater_equal(CBA_data["bcd"], parameters["selection_threshold"] * max_indicator), np.greater(CBA_data["bcd"], stop_criterion))
    else:
        print("No valid selection criterion")
        sys.exit()

    if not parameters["parallel_build"]: # Selection of more than one candidate is not usually allowed.
        for i, x in enumerate(selection_vector):
            if any(selection_vector[:i]):
                selection_vector[i] = False

    selected_expansions = CBA_data["valid expansions"].loc[selection_vector]

    # Transforms project combinations to single project expansions.
    for index, x in selected_expansions.iterrows():
        for y, z in zip(x[0], x[2]):
            selected_links.update({y: z})

    return selected_links, selection_vector, CBA_data


def Expansions_Iteration(period, selection_data, links_dataframe, pypsa_network):
    """ Calls the candidate portfolio creation, runs OPF for base case and all candidates and selects candidate"""

    # Declaration for base case variables.
    CBA_data = {}
    CBA_base = {}
    valid_grid = Network_Graph(period, pypsa_network, links_dataframe)

    CBA_base = {}

    global bus_columns, link_columns, gen_sto_columns

    bus_columns = pypsa_network.buses["number"]
    link_columns = links_dataframe.sort_values("number").index
    gen_sto_columns = list(pypsa_network.generators["name"]) + list(pypsa_network.storage_units["name"])

    index = pd.MultiIndex(levels=[[], []],
                          labels=[[], []],
                          names=[u'expansion', u'planning case'])

    CBA_base["valid expansions"] = pd.DataFrame(selection_data, columns=["Links", "igraph Links", "Capacity", "Buses", "Zones", "Type",
                                                                                  "Wind Sensitivity", "Capacity Sensitivity"])
    CBA_base["valid expansions"].index = ["base"]


    CBA_base["bus prices"] = pd.DataFrame(columns=bus_columns, index=index)
    CBA_base["nodal generation"] = CBA_base["bus prices"].copy()
    CBA_base["bus demand"] = CBA_base["bus prices"].copy()
    CBA_base["nodal gen cost"] = CBA_base["bus prices"].copy()
    CBA_base["marginal nodal cost"] = pd.DataFrame(columns=bus_columns, index=["expansion"])
    CBA_base["nodal sep inv cost"] = CBA_base["marginal nodal cost"].copy()
    CBA_base["dist nodal inv cost"] = CBA_base["marginal nodal cost"].copy()

    CBA_base["op cost"] = pd.DataFrame(index=index, columns=["op cost"])
    CBA_base["cost"] = pd.DataFrame(columns=["cost"])
    CBA_base["length"] = pd.DataFrame(columns=["length"])
    CBA_base["link flow"] = pd.DataFrame(columns=link_columns, index=index)
    CBA_base["link price delta"] = CBA_base["link flow"].copy()

    CBA_base["generator gen"] = pd.DataFrame(columns=gen_sto_columns, index=index)
    CBA_base["generator cost"] = CBA_base["generator gen"].copy()

    CBA_base["clusters"] = [None]

    Picture_Grid(period, valid_grid) # Draws base case grid.

    # Simulation of base case and CBA data calculation.
    if parameters["exp_detail"] >= 1:
        print("Running base case", period)

    pypsa_network.lines.loc[:, "s_nom"] = links_dataframe.loc[links_dataframe["link_type"] == "line", "s_nom"].mul(
        links_dataframe.loc[links_dataframe["link_type"] == "line", "active"]).values
    pypsa_network.transport_links.loc[:, "s_nom"] = links_dataframe.loc[links_dataframe["link_type"] == "link", "s_nom"].mul(
        links_dataframe.loc[links_dataframe["link_type"] == "link", "active"]).values

    output = Run_OPF(period, pypsa_network) # Base case OPF

    # Write_PyPSA(output[0])

    if False: # Print output for debugging.
        for x in output:
            Print_PyPSA(x)


    island_grid = valid_grid.copy()
    island_grid.delete_edges(island_grid.es.select(active=0))

    CBA_base["clusters"][0] = [island_grid.vs[y]["pandas_index"] for y in island_grid.clusters()] # Determines grid clusters for costs and benefits distribution.

    # Base case output recovery and CBA data calculation.
    CBA_base = Output_Recovery(CBA_base, 0, output, links_dataframe)
    CBA_base = Welfare_Calculation(CBA_base, pypsa_network)
    CBA_base = Onshore_Distribution(CBA_base, valid_grid, pypsa_network)

    # Simulation of candidates

    CBA_data = {}

    # Creation of candidate portfolio.
    valid_expansions = np.array(Expansion_Portfolio(period, CBA_base, pypsa_network, valid_grid))

    # Declaration for candidate variables.
    CBA_data["valid expansions"] = pd.DataFrame(valid_expansions,
                                                columns=["Links", "igraph Links", "Capacity", "Buses", "Zones", "Type",
                                                         "Wind Sensitivity", "Capacity Sensitivity"])

    CBA_data["valid expansions"]["Links String"] = CBA_data["valid expansions"]["Links"].astype(str)
    CBA_data["valid expansions"]["Capacity String"] = CBA_data["valid expansions"]["Capacity"].astype(str)

    # Elimination of eventual duplicate candidates before simulation
    CBA_data["valid expansions"] = CBA_data["valid expansions"].drop_duplicates(subset = ["Links String", "Capacity String"])
    parameters["number of expansions"] = len(CBA_data["valid expansions"].index)
    CBA_data["valid expansions"].index = range(parameters["number of expansions"])

    # Continue declaration for candidate variables.
    CBA_data["cost"] = pd.DataFrame(columns=["cost"], index=range(parameters["number of expansions"]))
    CBA_data["length"] = pd.DataFrame(columns=["length"])

    CBA_data["bus prices"] = pd.DataFrame(columns=bus_columns, index=index)
    CBA_data["nodal generation"] = CBA_data["bus prices"].copy()
    CBA_data["bus demand"] = CBA_data["bus prices"].copy()
    CBA_data["nodal gen cost"] = CBA_data["bus prices"].copy()


    CBA_data["marginal nodal cost"] = pd.DataFrame(columns=bus_columns)
    CBA_data["nodal inv cost"] = CBA_data["marginal nodal cost"].copy()
    CBA_data["nodal sep inv cost"] = CBA_data["marginal nodal cost"].copy()
    CBA_data["dist nodal inv cost"] = CBA_data["marginal nodal cost"].copy()
    CBA_data["on_bus_matrix"] = pd.DataFrame(False, index = range(parameters["number of expansions"]), columns=bus_columns)


    CBA_data["op cost"] = pd.DataFrame(index=index, columns=["op cost"])
    CBA_data["link flow"] = pd.DataFrame(columns=link_columns, index=index)
    CBA_data["link price delta"] = CBA_data["link flow"].copy()

    CBA_data["generator gen"] = pd.DataFrame(columns=gen_sto_columns, index=index)
    CBA_data["generator cost"] = CBA_data["generator gen"].copy()

    CBA_data["bilateral benefit"] = pd.Series(0, index=range(parameters["number of expansions"]))
    CBA_data["unilateral benefit"] = CBA_data["bilateral benefit"].copy()

    CBA_data["clusters"] = [None] * parameters["number of expansions"]

    # Iteration of simulation of candidates
    links_dataframe["s_nom_orig"] = links_dataframe["s_nom"]
    for i, [pandas_number, exp, cap, buses, zones, type, wind_sens, capacity_sens, links_string, capacity_string] in CBA_data["valid expansions"].iterrows():

        if parameters["exp_detail"] >= 1:
            print("Running", type, "expansion", i, ":", [valid_grid.es[y]['number'] for y in exp], cap)

        # Activation of candidate lines with capacity, reactance and resistance.
        transmission_status = [valid_grid.es[y]["active"] for y in exp]

        for y in range(len(exp)):
            link_number = str(valid_grid.es[exp[y]]["number"])  # keep as str to allow names

            if transmission_status[y] == 1:
                links_dataframe.loc[link_number, "r"] = links_dataframe.loc[link_number, "r"] * links_dataframe.loc[link_number, "s_nom_orig"] / (
                links_dataframe.loc[link_number, "s_nom_orig"] + cap[y])
                links_dataframe.loc[link_number, "x"] = links_dataframe.loc[link_number, "x"] * links_dataframe.loc[link_number, "s_nom_orig"] / (
                links_dataframe.loc[link_number, "s_nom_orig"] + cap[y])
                links_dataframe.loc[link_number, "s_nom"] += cap[y]
            else:
                valid_grid.es[exp[y]]["active"] = 1
                links_dataframe.loc[link_number, "active"] = 1
                links_dataframe.loc[link_number, "r"] = links_dataframe.loc[link_number, "r"] * links_dataframe.loc[link_number, "s_nom_orig"] / cap[y]
                links_dataframe.loc[link_number, "x"] = links_dataframe.loc[link_number, "x"] * links_dataframe.loc[link_number, "s_nom_orig"] / cap[y]
                links_dataframe.loc[link_number, "s_nom"] = cap[y]

        # Transfer of candidate line data to PyPSA.

        pypsa_network.lines.loc[:, "s_nom"] = links_dataframe.loc[links_dataframe["link_type"] == "line", "s_nom"].mul(
            links_dataframe.loc[links_dataframe["link_type"] == "line", "active"])
        pypsa_network.transport_links.loc[:, "s_nom"] = links_dataframe.loc[links_dataframe["link_type"] == "link", "s_nom"].mul(
            links_dataframe.loc[links_dataframe["link_type"] == "link", "active"])
        pypsa_network.lines.loc[:, ["r", "x"]] = links_dataframe.loc[links_dataframe["link_type"] == "line", ["r", "x"]]
        pypsa_network.transport_links.loc[:, ["r", "x"]] = links_dataframe.loc[links_dataframe["link_type"] == "link", ["r", "x"]]

        output = Run_OPF(period, pypsa_network) # Candidate OPF

        # Output recovery per candidate.
        CBA_data = Output_Recovery(CBA_data, i, output, links_dataframe)

        island_grid = valid_grid.copy()
        island_grid.delete_edges(island_grid.es.select(active=0))
        CBA_data["clusters"][i] = [island_grid.vs[y]["pandas_index"] for y in island_grid.clusters()] # Determines grid clusters for costs and benefits distribution.

        # Deactivation of candidate lines with capacity, reactance and resistance.
        for y in range(len(exp)):
            link_number = str(valid_grid.es[exp[y]]["number"])

            if transmission_status[y] == 1:
                links_dataframe.loc[link_number, "r"] = links_dataframe.loc[link_number, "r"] * (links_dataframe.loc[link_number, "s_nom_orig"] + cap[y]) / links_dataframe.loc[
                    link_number, "s_nom_orig"]
                links_dataframe.loc[link_number, "x"] = links_dataframe.loc[link_number, "x"] * (links_dataframe.loc[link_number, "s_nom_orig"] + cap[y]) / links_dataframe.loc[
                    link_number, "s_nom_orig"]
                links_dataframe.loc[link_number, "s_nom"] -= cap[y]

            else:
                valid_grid.es[exp[y]]["active"] = 0
                links_dataframe.loc[link_number, "active"] = 0
                links_dataframe.loc[link_number, "r"] = links_dataframe.loc[link_number, "r"] * cap[y] / links_dataframe.loc[link_number, "s_nom_orig"]
                links_dataframe.loc[link_number, "x"] = links_dataframe.loc[link_number, "x"] * cap[y] / links_dataframe.loc[link_number, "s_nom_orig"]
                links_dataframe.loc[link_number, "s_nom"] = links_dataframe.loc[link_number, "s_nom_orig"]

    # Simultaneous CBA data calculation for all candidates.
    CBA_data = Cost_Calculation(CBA_data, valid_grid, pypsa_network)
    CBA_data = Welfare_Calculation(CBA_data, pypsa_network, 1)
    CBA_data = Onshore_Distribution(CBA_data, valid_grid, pypsa_network, 1)
    CBA_data = Cost_Distribution(CBA_data, CBA_base)

    # Candidate selection.
    selected_links, selection_vector, CBA_data = Expansion_Selection(CBA_base, CBA_data)

    # Creation of dataframe for output printing.
    CBA_data["valid expansions"].loc[:, "consumer payments"] = CBA_data["consumer payments"]
    CBA_data["valid expansions"].loc[:, "producer surplus"] = CBA_data["producer surplus"]
    CBA_data["valid expansions"].loc[:, "congestion rent"] = CBA_data["congestion rent"]
    CBA_data["valid expansions"].loc[:, "op cost"] = CBA_data["total op cost"]
    CBA_data["valid expansions"].loc[:, "bcr"] = CBA_data["bcr"]
    CBA_data["valid expansions"].loc[:, "cost"] = CBA_data["cost"]
    CBA_data["valid expansions"].loc[:, "social cost"] = CBA_data["net social cost benefit"]
    CBA_data["valid expansions"].loc[:, "social benefit"] = CBA_data["dist net nodal benefit"].sum(axis=1)
    CBA_data["valid expansions"].loc[:, "bilateral benefit"] = CBA_data["bilateral benefit"]
    CBA_data["valid expansions"].loc[:, "unilateral benefit"] = CBA_data["unilateral benefit"]
    CBA_data["valid expansions"].loc[:, "Selection"] = selection_vector

    CBA_base["valid expansions"].loc["base", "consumer payments"] = CBA_base["consumer payments"].loc[0]
    CBA_base["valid expansions"].loc["base", "producer surplus"] = CBA_base["producer surplus"].loc[0]
    CBA_base["valid expansions"].loc["base", "congestion rent"] = CBA_base["congestion rent"].loc[0]
    CBA_base["valid expansions"].loc["base", "op cost"] = CBA_base["total op cost"].loc[0]

    # Selected candidate data.
    selection_data = CBA_data["valid expansions"][CBA_data["valid expansions"]["Selection"] == True]

    if len(selection_data) == 0:
        selection_data = Reset_Selection_Data(pypsa_network) # If no candidate was selected in period correct selection_data.

    # Print period output.
    if True:
        CBA_print = pd.concat([CBA_base["valid expansions"], CBA_data["valid expansions"]])

        print(CBA_print.loc[:, ["Links", "Type", "Capacity", "Selection", "consumer payments", "producer surplus", "congestion rent", "op cost", "cost", "bcr"]])

        print("Selected expansion distributed net benefits")

        print(CBA_data["dist net nodal benefit"].loc[selection_vector,:])

    return selected_links, selection_data, valid_grid, pypsa_network, links_dataframe, CBA_data, period


def Print_PyPSA(output_network):
    """ Prints OPF results for debugging """

    output_network.generators["generation"] = output_network.generators_t.p.values[0]

    print("\nCost\n", output_network.objective)
    print("\nBuses Balance\n", output_network.buses_t.p)
    print("\nBuses Demand\n", output_network.loads_t.p)
    print("\nBuses Price\n", output_network.buses_t.marginal_price)
    print("\nGenerators\n", output_network.generators.loc[:, ["name", "bus", "generation", "source", "p_nom", "marginal_cost", "p_nom_series"]].sort(columns=["bus", "name"]))
    print("\nGeneration\n", output_network.generators_t.p)
    print("\nGeneration Max P\n", output_network.generators_t.p_max_pu)
    print("\nStorage Units\n", output_network.storage_units)
    print("\nStorage Discharge\n", output_network.storage_units_t.p)
    print("\nState of Charge\n", output_network.storage_units_t.state_of_charge)
    print("\nInflow\n", output_network.storage_units_t.inflow)
    print("\nLines\n", output_network.lines)
    print("\nLines Use\n", output_network.lines_t.p0)
    print("\nTransport Links\n", output_network.transport_links)
    print("\nTransport Links Use\n", output_network.transport_links_t.p0)
    print("\nDemand \n", output_network.loads_t.p_set)

    return


def Write_PyPSA(output_network):
    """ Writes OPF results for debugging """

    output_network.export_to_csv_folder(os.path.join(os.getcwd(), r"pypsa_network\results"),
                                        time_series={"generators": {"p_max_pu": None, "p": None},
                                                     "loads": {"p_set": None}, "lines": {"p1": None, "p0": None},
                                                     "transport_links": {"p1": None, "p0": None},
                                                     "buses": {"marginal_price": None, "p": None},
                                                     "storage_units": {"inflow": None, "p_max_pu": None, "p": None, "state_of_charge": None,
                                                                       "state_of_charge_set": None, "spill": None}},
                                        verbose=False)
    return

def Write_Output(CBA_data, period, run_number):
    """ Writes output with all candidates in the period in CSV format for specific run"""

    writecsv = open(r"OGEM_data"+str(run_number)+".csv", "a", newline='')
    writer = csv.writer(writecsv, delimiter=';', quoting=csv.QUOTE_NONE)
    for i, x in CBA_data["valid expansions"].iterrows():
        writer.writerow(
            [parameters["simulation_name"], parameters["run_number"], parameters["periods"],
             parameters["planning_cases"], str(parameters["sens_name"]), str(parameters["current_parameters_value"]),
             period, str(i), str(x["Links"]), str(x["Type"]), str(x["Capacity"]), x["Wind Sensitivity"], x["Capacity Sensitivity"],
             str(x["Selection"]), CBA_data["bcr"].loc[i],
             CBA_data["dist net nodal benefit"].loc[i, :].sum(), x.loc["bilateral benefit"],
             x.loc["unilateral benefit"], CBA_data["cost"].loc[i, "cost"], parameters["on_tc"],
             parameters["off_tc"], parameters["owf_tc"], parameters["cap_length"], parameters["cap"],
             parameters["length"], parameters["multiterminal_costs"],
             parameters["selection_criterion"], parameters["benefit_type"]] + [j for j in CBA_data["nodal prices"].loc[i, :]])
    writecsv.close()

    return

def Create_Output_Files(f):
    """" Creates excel file with run parameters and csv file where output of all runs is aggregated """

    write_workbook = xls.workbook.Workbook()

    for file in f:
        if file[:4] == "OGEM":
            simulation_file = os.path.join("Simulations", file)
            ws1 = write_workbook.create_sheet(title=file)
            warnings.simplefilter("ignore")
            simulations_workbook = xls.load_workbook(simulation_file, data_only=True)
            warnings.simplefilter("default")

            for row in simulations_workbook["parameters"].rows:
                for cell in row:
                    ws1[cell.coordinate] = cell.value

    write_workbook.save("OGEM_setup_data.xlsx")

    writecsv = open(r"OGEM_data.csv", "w", newline='')
    writer = csv.writer(writecsv, delimiter=';', quoting=csv.QUOTE_NONE)
    writer.writerow(["Simulation Name", "Run Number", "Periods", "Planning Cases", "Sens. Parameter", "Value", "Period",
                     "Expansion Number", "Expansion", "Type",
                     "Capacity", "Wind Multiplier", "Capacity Multiplier", "Selected", "BCR", "Global", "Bilateral",
                     "Unilateral", "Cost", "con", "cowh", "cowf", "clcl", "clc", "cll", "multiterminal costs",
                     "Selection Criterion", "Benefit Type"] + [x for x in range(parameters["buses"])])
    writecsv.close()

    return

def Join_XLS(number_runs):
    """ Joins output of all runs in existing csv file """

    writecsv = open(r"OGEM_data.csv", "a", newline='')
    writer = csv.writer(writecsv, delimiter=';',quoting=csv.QUOTE_NONE)

    for run in range(number_runs):
        readcsv = open(r"OGEM_data"+str(run)+".csv", "r", newline='')
        reader = csv.reader(readcsv, delimiter=';',quoting=csv.QUOTE_NONE)
        for row in reader:
            writer.writerow(row)
        readcsv.close()
    writecsv.close()

    for run in range(number_runs):
        if os.path.isfile(r"OGEM_data" + str(run) + ".csv"):
            os.remove(r"OGEM_data" + str(run) + ".csv")
    return


def PyPSA_Network_Setup(period, pypsa_network):
    """ Creation of PyPSA elements from input data """

    # PyPSA snapshot data: name, weightings.
    columns = ["name", "weightings"]
    index = list(range(parameters["case_segments"]))
    pypsa_network.set_snapshots(index)
    pypsa_network.snapshot_weightings = parameters["segments_probability"]

    # PyPSA bus data: name, x, y, terminal_type, shared_tso, tso, active, current_type.
    columns = ["number", "name", "valid", "terminal_type", "x", "y", "shared_tso", "current_type", "zone", "offshore_p_nom"]
    index = [str(x[0]) for x in OGEM_settings.data["buses"]]
    data = [x[0:8] + [1, 0] for x in OGEM_settings.data["buses"]] # Initial TSO zone and offshore power are 1 and 0.
    dataframe = pd.DataFrame(data, columns=columns, index=index)
    dataframe["shared_tso"] = [list(x.split()) for x in dataframe["shared_tso"]]
    dataframe["tso"] = dataframe["shared_tso"]

    pypsa.io.import_components_from_dataframe(pypsa_network, dataframe, "Bus")
    pypsa_network.buses["number"] = pypsa_network.buses["number"].astype(int)

    # PyPSA line data: name, bus0, bus1, x, r, s_nom, s_nom_extendable, length.
    # PyPSA transport link data: name, bus0, bus1, s_nom, s_nom_extendable.
    columns = ["number", "bus0", "bus1", "name", "active", "s_nom", "length", "link_type", "valid"]
    index = [str(x[0]) for x in OGEM_settings.data["transmission"]]
    data = [x[0:10] for x in OGEM_settings.data["transmission"]]

    # links_dataframe is used to update grid data for each candidate expansion and then transfer this to the PyPSA dataframe in the candidate iteration.
    links_dataframe = pd.DataFrame(data, columns=columns,
                                   index=index)
    links_dataframe["r"] = links_dataframe["length"] * 0.01 * OGEM_settings.base_Z / 1.6
    links_dataframe["x"] = links_dataframe["r"]
    links_dataframe["s_nom_extendable"] = False
    links_dataframe = links_dataframe.sort_values("link_type")

    pypsa.io.import_components_from_dataframe(pypsa_network, links_dataframe[links_dataframe["link_type"] == "line"], "Line")
    pypsa.io.import_components_from_dataframe(pypsa_network, links_dataframe[links_dataframe["link_type"] == "link"],
                                              "TransportLink")

    # PyPSA indexes are str so create a number index as int for igraph and numpy indexing.
    links_dataframe["number"] = links_dataframe["number"].astype(int)
    pypsa_network.lines["number"] = pypsa_network.lines["number"].astype(int)
    pypsa_network.transport_links["number"] = pypsa_network.transport_links["number"].astype(int)

    # PyPSA generator data: name, bus, dispatch, p_nom, p_nom_extendable, source, marginal_cost.
    columns = ["name", "source", "bus", "p_nom", "marginal_cost"]
    index = [str(x[0]) for x in OGEM_settings.data["generation"]]
    data = [x[0:5] for x in OGEM_settings.data["generation"]]
    dataframe = pd.DataFrame(data, columns=columns,
                             index=index)

    dataframe["p_nom_extendable"] = False
    dataframe["p_nom_series"] = [list(map(float, x.split())) for x in dataframe["p_nom"].astype(str)]
    dataframe["p_nom"] = [x[period] for x in dataframe["p_nom_series"]]

    # Variable technologies are wind (offshore and onshore) and PV
    dataframe["dispatch"] = ["variable" if x in ["off_wind", "PV", "on_wind"] else "flexible" for x in dataframe["source"]]

    pypsa.io.import_components_from_dataframe(pypsa_network, dataframe, "Generator")

    pypsa_network.buses["p_nom"] = pypsa_network.generators.loc[:, ["bus", "p_nom"]].groupby(by="bus").sum()

    # PyPSA storage data: name, bus, dispatch, p_nom, p_nom_extendable, source, state_of_charge_initial, efficiency_store, efficiency_dispatch, marginal_cost
    columns = ["name", "source", "bus", "p_nom", "efficiency_store", "efficiency_dispatch", "state_of_charge_initial", "marginal_cost"]

    index = [x[1] for x in OGEM_settings.data["storage"]]
    data = [x[1:9] for x in OGEM_settings.data["storage"]]
    dataframe = pd.DataFrame(data, columns=columns,
                             index=index)

    dataframe["dispatch"] = "flexible"
    dataframe["p_nom_extendable"] = False

    dataframe["p_nom_series"] = [list(map(float, x.split())) for x in dataframe["p_nom"]]
    dataframe["p_nom"] = [max(x) for x in dataframe["p_nom_series"]]

    pypsa.io.import_components_from_dataframe(pypsa_network, dataframe, "StorageUnit")

    # PyPSA load data: bus, name
    columns = ["bus", "name"]
    index = pypsa_network.buses.index
    data = pypsa_network.buses[["number", "name"]].rename(columns={"number": "bus"})

    dataframe = pd.DataFrame(data, index=index, columns=columns)

    dataframe["bus"] = dataframe["bus"].astype(int)

    pypsa.io.import_components_from_dataframe(pypsa_network, dataframe, "Load")

    return links_dataframe, pypsa_network


def PyPSA_Network_Update(period, pypsa_network, links_dataframe):
    """ Updates PyPSA data with period changes """

    # Copy expanded grid from links_dataframe to PyPSA with capacity, status, reactance and resistance.
    pypsa_network.lines["x"] = links_dataframe.loc[links_dataframe["link_type"] == "line", "x"]
    pypsa_network.lines["r"] = links_dataframe.loc[links_dataframe["link_type"] == "line", "r"]
    pypsa_network.lines["s_nom"] = links_dataframe.loc[links_dataframe["link_type"] == "line", "s_nom"]
    pypsa_network.lines["active"] = links_dataframe.loc[links_dataframe["link_type"] == "line", "active"]

    pypsa_network.transport_links["x"] = links_dataframe.loc[links_dataframe["link_type"] == "link", "x"]
    pypsa_network.transport_links["r"] = links_dataframe.loc[links_dataframe["link_type"] == "link", "r"]
    pypsa_network.transport_links["s_nom"] = links_dataframe.loc[links_dataframe["link_type"] == "link", "s_nom"]
    pypsa_network.transport_links["active"] = links_dataframe.loc[links_dataframe["link_type"] == "link", "active"]

    # Update time series for generation, storage
    pypsa_network.generators.loc[:, "p_nom"] = [x[period] for x in pypsa_network.generators["p_nom_series"]]

    pypsa_network.storage_units.loc[:, "p_nom"] = [x[period] for x in pypsa_network.storage_units["p_nom_series"]]

    # p_nom_offshore and previous_p_nom are used to define newly added capacity in the candidate portfolio creation.
    p_nom_offshore = pypsa_network.generators.loc[pypsa_network.generators["source"] == "off_wind", ["bus", "p_nom"]].groupby(by="bus").sum()

    pypsa_network.buses["p_nom_offshore"] = p_nom_offshore.rename(columns={"p_nom": "p_nom_offshore"})
    if period != 0:
        pypsa_network.buses["previous_p_nom"] = pypsa_network.buses["p_nom"]
    else:
        pypsa_network.buses["previous_p_nom"] = 0

    pypsa_network.buses["p_nom"] = pypsa_network.generators.loc[:, ["bus", "p_nom"]].groupby(by="bus").sum()

    return links_dataframe, pypsa_network
    

def PyPSA_Time_Series(data, period, case, network):
    """ Updates times series for the current period """
    # PyPSA load time series: snapshots, loads
    columns = network.loads.index
    index = network.snapshots
    demand_series = np.array([[bus_demand[4] for bus_demand in data["demand_series"] if (bus_demand[2] == period and bus_demand[3] == case)] for segment in index])
    dataframe = pd.DataFrame(demand_series, columns=columns, index=index)
    pypsa.io.import_series_from_dataframe(network, dataframe, "Load", "p_set")

    # PyPSA RES generator time series: snapshots, generators
    index = network.snapshots
    columns = [str(RES_line[8]) for RES_line in data["RES_series"] if RES_line[3] == case and RES_line[4] == 0]
    RES_series = np.array(
        [[RES_line[6] for RES_line in data["RES_series"] if RES_line[3] == case and RES_line[4] == segment] for segment
         in range(parameters["case_segments"])])
    dataframe = pd.DataFrame(RES_series, columns=columns, index=index)

    pypsa.io.import_series_from_dataframe(network, dataframe, "Generator", "p_max_pu")

    if True: # Inflow storage technologies need their inflow time series.
        # PyPSA hydro generator time series: snapshots, generators
        index = network.snapshots
        columns = [str(hydro_line[8]) for hydro_line in data["inflow_series"] if hydro_line[3] == case and hydro_line[4] == 0]
        hydro_series = np.array(
            [[hydro_line[6] * network.storage_units.loc[str(hydro_line[8]), "p_nom"] for hydro_line in data["inflow_series"] if hydro_line[3] == case and hydro_line[4] == segment]
             for segment in range(parameters["case_segments"])])

        dataframe = pd.DataFrame(hydro_series, columns=columns, index=index)

        pypsa.io.import_series_from_dataframe(network, dataframe, "StorageUnit", "inflow")

    return network


def Load_Parameters(simulation_file):
    """ Load analysis parameters to initialize run calls """

    global parameters
    warnings.simplefilter("ignore")
    simulations_workbook = xls.load_workbook(simulation_file, data_only=True)
    warnings.simplefilter("default")

    parameters = {}

    main_raw = [[y.value for y in x] for x in simulations_workbook["parameters"].rows]
    parameters.update({x[0]: x[1] for x in main_raw[1:] if x[0] != None})
    parameters["sens_name"] = parameters["sens_name"].split() # Name of the analysis parameters.
    parameters["sens_values"] = [map(float, x.split()) for x in str(parameters["sens_value"]).split(';')] # Value of the parameters corresponding to the names.
    parameters["segments_probability"] = [float(x) for x in str(parameters["segments_probability"]).split()]
    parameters["cases_probability"] = [float(x) for x in str(parameters["cases_probability"]).split()]

    return parameters

def Load_Data(simulation_file):
    """ Load input data for a given run """
    warnings.simplefilter("ignore")
    simulations_workbook = xls.load_workbook(simulation_file, data_only=True)
    warnings.simplefilter("default")

    data = {}

    main_raw = [[y.value for y in x] for x in simulations_workbook["parameters"].rows]

    transmission_raw = [[y.value for y in x] for x in simulations_workbook["transmission"].rows]
    data["transmission"] = [x for x in transmission_raw[1:] if x[0] != None]

    buses_raw = [[y.value for y in x] for x in simulations_workbook["buses"].rows]
    data["buses"] = [x for x in buses_raw[1:] if x[2]]

    generation_raw = [[y.value for y in x] for x in simulations_workbook["generation"].rows]
    data["generation"] = [x for x in generation_raw[1:] if x[1] != None and x[3] != 0]

    storage_raw = [[y.value for y in x] for x in simulations_workbook["storage"].rows]
    data["storage"] = [x for x in storage_raw[1:] if x[2] != None and x[4] != 0]

    inflow_raw = [[y.value for y in x] for x in simulations_workbook["inflow_series"].rows]
    data["inflow_series"] = [x for x in inflow_raw[1:] if x[1] != None and x[5] != 0]

    RES_raw = [[y.value for y in x] for x in simulations_workbook["RES_series"].rows]
    data["RES_series"] = [x for x in RES_raw[1:] if x[1] != None and x[5] != 0]

    demand_raw = [[y.value for y in x] for x in simulations_workbook["demand_series"].rows]
    data["demand_series"] = [x for x in demand_raw[1:] if x[4] >= 0]

    res_availability_raw = [[y.value for y in x] for x in simulations_workbook["RES_availability"].rows]
    data["RES_availability"] = [x for x in res_availability_raw if x[3] != None]

    return data


def Reset_Selection_Data(pypsa_network):
    """ Resets selection data for first period or in case no candidate is selected """
    selected_expansion_data = [[(), (), [], [], {}, "None", 1.0, 1.0]]

    selected_expansion_data[0][3] = [x["number"] for i, x in pypsa_network.buses.iterrows() if x["terminal_type"] in ["on", "on ex", "on im"]]

    return selected_expansion_data

